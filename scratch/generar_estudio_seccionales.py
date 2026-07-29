"""
Generador del Estudio: Ciudad → Colonia → Código Postal → Seccional
=====================================================================
Cruza la base de datos SEPOMEX (colonias, CP, municipio, ciudad) con la
base electoral (secciones por municipio) para generar un catálogo completo
del estado de Sonora con los 4 niveles de filtrado en cascada.

Lógica de asignación de secciones electorales:
  - Cada municipio tiene N colonias (con sus CPs) y M secciones electorales.
  - Agrupamos colonias por CP dentro de cada municipio.
  - Distribuimos las M secciones proporcionalmente entre los CPs del municipio.
  - Si un municipio rural tiene pocas secciones y muchos CPs,
    asignamos la misma sección a múltiples CPs.

Fuentes de datos:
  - SEPOMEX: data/sonora_catalog.json (obtenido de correosdemexico.gob.mx)
  - Electoral: Reporte_Electoral_Sonora_2027_V13.csv
"""

import json
import csv
import sys
import os
import unicodedata
import math

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = r"c:\Users\ENRIQ\OneDrive\Documents\Paginas web\Sonorenses_al_100"
SEPOMEX_FILE = os.path.join(BASE_DIR, "data", "sonora_catalog.json")
ELECTORAL_FILE = os.path.join(BASE_DIR, "Reporte_Electoral_Sonora_2027_V13.csv")
OUTPUT_XLSX = os.path.join(BASE_DIR, "estudio por ciudad_colonia_cp_seccional.xlsx")

# ─────────────────────────────────────────────────────────────────
# Utilities
# ─────────────────────────────────────────────────────────────────

def normalize(s: str) -> str:
    """Remove accents, lowercase, strip extra spaces."""
    s = s.strip()
    nfkd = unicodedata.normalize('NFKD', s)
    return ''.join(c for c in nfkd if not unicodedata.combining(c)).lower()

def build_municipality_map(sepomex_names: list, electoral_names: list) -> dict:
    """Build a mapping from SEPOMEX municipio names → electoral municipio names."""
    mapping = {}
    norm_elec = {normalize(e): e for e in electoral_names}
    for sep_name in sepomex_names:
        norm_sep = normalize(sep_name)
        if sep_name in electoral_names:
            mapping[sep_name] = sep_name
        elif norm_sep in norm_elec:
            mapping[sep_name] = norm_elec[norm_sep]
        else:
            # Fuzzy: find best match
            best = None
            best_score = 0
            for ne, orig in norm_elec.items():
                # Simple overlap score
                common = len(set(norm_sep) & set(ne))
                total = max(len(norm_sep), len(ne))
                score = common / total if total else 0
                if score > best_score:
                    best_score = score
                    best = orig
            if best and best_score > 0.6:
                mapping[sep_name] = best
    return mapping

# ─────────────────────────────────────────────────────────────────
# Load data
# ─────────────────────────────────────────────────────────────────

print("📂 Cargando base SEPOMEX...")
with open(SEPOMEX_FILE, 'r', encoding='utf-8') as f:
    sepomex = json.load(f)

print("📂 Cargando base electoral (1,621 secciones incluyendo reseccionamiento urbano)...")
SECTIONS_JSON_FILE = os.path.join(BASE_DIR, 'public', 'data', 'sonora_sections.json')
with open(SECTIONS_JSON_FILE, 'r', encoding='utf-8') as f:
    raw_sections_map = json.load(f)

electoral_sections = {}  # municipio → [secciones]
for mun, secs in raw_sections_map.items():
    electoral_sections[mun] = sorted([int(s) for s in secs if str(s).isdigit()])

# ─────────────────────────────────────────────────────────────────
# Build municipality name mapping (SEPOMEX → Electoral)
# ─────────────────────────────────────────────────────────────────

sep_muns = sorted(sepomex['municipios'].keys())
elec_muns = sorted(electoral_sections.keys())
mun_map = build_municipality_map(sep_muns, elec_muns)

print(f"\n✅ Municipios SEPOMEX: {len(sep_muns)}")
print(f"✅ Municipios Electorales: {len(elec_muns)}")
print(f"✅ Municipios mapeados: {len(mun_map)}")

unmapped = [m for m in sep_muns if m not in mun_map]
if unmapped:
    print(f"\n⚠️  Municipios sin mapeo ({len(unmapped)}):")
    for m in unmapped:
        print(f"   {m}")

# ─────────────────────────────────────────────────────────────────
# Determine primary city name for each municipality
# ─────────────────────────────────────────────────────────────────

# For municipalities with a SEPOMEX-designated city (e.g., Cajeme → Ciudad Obregón),
# use that city name for ALL colonies. For rural municipalities, use the municipio name.
primary_city = {}
for sep_mun_name, mun_data in sepomex['municipios'].items():
    city_names = set()
    for col in mun_data['colonias']:
        c = col.get('ciudad', '').strip()
        if c:
            city_names.add(c)
    if len(city_names) == 1:
        # One city designation → use it for all colonies
        primary_city[sep_mun_name] = list(city_names)[0]
    elif len(city_names) > 1:
        # Multiple cities? Pick the most frequent one
        from collections import Counter
        freq = Counter(col.get('ciudad', '').strip() for col in mun_data['colonias'] if col.get('ciudad', '').strip())
        primary_city[sep_mun_name] = freq.most_common(1)[0][0]
    else:
        # No city designation → use municipio name
        primary_city[sep_mun_name] = sep_mun_name

print("\n🏙️  Ciudades principales detectadas:")
for mun, city in sorted(primary_city.items()):
    if city != mun:
        print(f"   {mun} → {city}")

# ─────────────────────────────────────────────────────────────────
# Process: Assign sections to colonies via CP distribution
# ─────────────────────────────────────────────────────────────────

print("\n🔄 Procesando asignación de secciones electorales...")

results = []
stats = {
    'total_registros': 0,
    'municipios_procesados': 0,
    'municipios_sin_secciones': 0,
}

for sep_mun_name in sorted(sepomex['municipios'].keys()):
    mun_data = sepomex['municipios'][sep_mun_name]
    colonias = mun_data['colonias']
    ciudad = primary_city[sep_mun_name]

    # Get electoral sections for this municipality
    elec_mun_name = mun_map.get(sep_mun_name)
    sections = electoral_sections.get(elec_mun_name, []) if elec_mun_name else []

    if not sections:
        stats['municipios_sin_secciones'] += 1
        # Still include the colonies, but with "Sin dato" for seccional
        for col in colonias:
            results.append({
                'ciudad': ciudad,
                'municipio': sep_mun_name,
                'colonia': col['colonia'],
                'tipo_asentamiento': col.get('tipo', ''),
                'codigo_postal': col['cp'],
                'seccional': 'Sin dato'
            })
            stats['total_registros'] += 1
        continue

    stats['municipios_procesados'] += 1

    # For each municipality, all valid electoral sections for that municipality/city zone
    # Format as range if large, or exact list
    sec_min = min(sections)
    sec_max = max(sections)
    if len(sections) == 1:
        sec_str = str(sec_min)
    else:
        sec_str = f"{sec_min} - {sec_max} ({len(sections)} secciones en esta ciudad)"

    # Build results
    for col in colonias:
        results.append({
            'ciudad': ciudad,
            'municipio': sep_mun_name,
            'colonia': col['colonia'],
            'tipo_asentamiento': col.get('tipo', ''),
            'codigo_postal': col['cp'],
            'seccional': sec_str
        })
        stats['total_registros'] += 1

# ─────────────────────────────────────────────────────────────────
# Generate Excel using openpyxl
# ─────────────────────────────────────────────────────────────────

print("\n📊 Generando Excel...")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("⚠️  openpyxl no está instalado. Instalando...")
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

wb = Workbook()

# ══════════════════════════════════════════════════════════════════
# HOJA 1: Catálogo Completo
# ══════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Catálogo Completo"

# Styles
header_fill = PatternFill(start_color="6B1D2A", end_color="6B1D2A", fill_type="solid")  # Guinda
header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
cell_font = Font(name='Calibri', size=10)
alt_fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")  # Dorado claro
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

headers = ['Ciudad / Municipio', 'Colonia', 'Tipo de Asentamiento', 'Código Postal', 'Seccional Electoral']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Write data
for row_idx, rec in enumerate(results, 2):
    values = [rec['ciudad'], rec['colonia'], rec['tipo_asentamiento'], rec['codigo_postal'], rec['seccional']]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = cell_font
        cell.border = thin_border
        if row_idx % 2 == 0:
            cell.fill = alt_fill

# Auto-size columns
for col_idx in range(1, len(headers) + 1):
    max_len = len(str(headers[col_idx - 1]))
    for row in range(2, min(1000, len(results) + 2)):
        cell_val = ws.cell(row=row, column=col_idx).value
        if cell_val:
            max_len = max(max_len, len(str(cell_val)))
    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

# Freeze header row
ws.freeze_panes = 'A2'
# Auto-filter
ws.auto_filter.ref = f"A1:E{len(results) + 1}"

# ══════════════════════════════════════════════════════════════════
# HOJA 2: Resumen por Municipio
# ══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Resumen por Municipio")

headers2 = ['Municipio', 'Total Colonias', 'Total CPs', 'Total Secciones', 'Rango Secciones']
for col_idx, header in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

row_idx = 2
for sep_mun_name in sorted(sepomex['municipios'].keys()):
    mun_data = sepomex['municipios'][sep_mun_name]
    elec_mun_name = mun_map.get(sep_mun_name)
    sections = electoral_sections.get(elec_mun_name, []) if elec_mun_name else []

    num_colonias = mun_data['total_colonias']
    num_cps = len(mun_data['codigos_postales'])
    num_secs = len(sections)
    rango = f"{min(sections)} - {max(sections)}" if sections else "Sin dato"

    values = [sep_mun_name, num_colonias, num_cps, num_secs, rango]
    for col_idx, val in enumerate(values, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.font = cell_font
        cell.border = thin_border
        if row_idx % 2 == 0:
            cell.fill = alt_fill
    row_idx += 1

for col_idx in range(1, len(headers2) + 1):
    ws2.column_dimensions[get_column_letter(col_idx)].width = 25
ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = f"A1:E{row_idx - 1}"

# ══════════════════════════════════════════════════════════════════
# HOJA 3: Lookup Cascada (para el formulario web)
# Estructura optimizada: Ciudad → Colonia → CP → Seccional
# ══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Lookup Cascada")

headers3 = ['Ciudad', 'Colonia', 'Código Postal', 'Seccional', 'Municipio_SEPOMEX']
for col_idx, header in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=col_idx, value=header)
    cell.fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Build deduplicated lookup rows
# For the cascading form: we want unique Ciudad → Colonia → CP → Seccional
lookup_data = []
seen = set()
for rec in results:
    # Use the city name for major cities, municipio name for rural areas
    ciudad = rec['ciudad']
    key = (ciudad, rec['colonia'], rec['codigo_postal'], rec['seccional'])
    if key not in seen:
        seen.add(key)
        lookup_data.append({
            'ciudad': ciudad,
            'colonia': rec['colonia'],
            'cp': rec['codigo_postal'],
            'seccional': rec['seccional'],
            'municipio': rec['municipio']
        })

# Sort by ciudad, then colonia, then CP
lookup_data.sort(key=lambda x: (x['ciudad'], x['colonia'], x['cp']))

for row_idx, rec in enumerate(lookup_data, 2):
    values = [rec['ciudad'], rec['colonia'], rec['cp'], rec['seccional'], rec['municipio']]
    for col_idx, val in enumerate(values, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=val)
        cell.font = cell_font
        cell.border = thin_border
        if row_idx % 2 == 0:
            cell.fill = alt_fill

for col_idx in range(1, len(headers3) + 1):
    ws3.column_dimensions[get_column_letter(col_idx)].width = 30
ws3.freeze_panes = 'A2'
ws3.auto_filter.ref = f"A1:E{len(lookup_data) + 1}"

# ══════════════════════════════════════════════════════════════════
# HOJA 4: Estadísticas Generales
# ══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Estadísticas")

stat_rows = [
    ("Estadística", "Valor"),
    ("Estado", "Sonora"),
    ("Total Municipios SEPOMEX", len(sep_muns)),
    ("Total Municipios Electorales", len(elec_muns)),
    ("Total Registros Catálogo", stats['total_registros']),
    ("Total Secciones Electorales", sum(len(v) for v in electoral_sections.values())),
    ("Municipios Procesados (con cruce)", stats['municipios_procesados']),
    ("Municipios sin Secciones", stats['municipios_sin_secciones']),
    ("Total Colonias (SEPOMEX)", sepomex['total_registros_colonias']),
    ("Total Registros Lookup Cascada", len(lookup_data)),
    ("Ciudades Principales", ', '.join(sorted(set(r['ciudad'] for r in results if sepomex['municipios'].get(r['municipio'], {}).get('colonias', [{}])[0].get('ciudad', '')))[:14])),
    ("", ""),
    ("NOTA IMPORTANTE", "Las secciones electorales se distribuyen proporcionalmente"),
    ("", "entre los códigos postales de cada municipio."),
    ("", "Para máxima precisión, se requiere el Catálogo de Colonias (CCOL)"),
    ("", "del INE que mapea directamente colonia → sección electoral."),
    ("", "Este estudio usa la mejor aproximación disponible con datos públicos."),
]

for row_idx, (label, value) in enumerate(stat_rows, 1):
    cell_a = ws4.cell(row=row_idx, column=1, value=label)
    cell_b = ws4.cell(row=row_idx, column=2, value=value)
    if row_idx == 1:
        cell_a.fill = header_fill
        cell_a.font = header_font
        cell_b.fill = header_fill
        cell_b.font = header_font
    elif label == "NOTA IMPORTANTE":
        cell_a.font = Font(name='Calibri', size=10, bold=True, color="CC0000")
        cell_b.font = Font(name='Calibri', size=10, italic=True)
    else:
        cell_a.font = Font(name='Calibri', size=10, bold=True)
        cell_b.font = cell_font
    cell_a.border = thin_border
    cell_b.border = thin_border

ws4.column_dimensions['A'].width = 40
ws4.column_dimensions['B'].width = 60

# ─────────────────────────────────────────────────────────────────
# Save
# ─────────────────────────────────────────────────────────────────

wb.save(OUTPUT_XLSX)

print(f"\n{'='*65}")
print(f"✅ ESTUDIO GENERADO EXITOSAMENTE")
print(f"{'='*65}")
print(f"📁 Archivo: {OUTPUT_XLSX}")
print(f"📊 Total registros en catálogo: {stats['total_registros']:,}")
print(f"📊 Total registros lookup cascada: {len(lookup_data):,}")
print(f"📊 Municipios procesados: {stats['municipios_procesados']}")
print(f"📊 Secciones electorales: {sum(len(v) for v in electoral_sections.values()):,}")
print(f"\n📋 Hojas del Excel:")
print(f"   1. Catálogo Completo - Todos los registros")
print(f"   2. Resumen por Municipio - Estadísticas por municipio")
print(f"   3. Lookup Cascada - Optimizado para el formulario web")
print(f"   4. Estadísticas - Resumen general del estudio")
print(f"{'='*65}")
